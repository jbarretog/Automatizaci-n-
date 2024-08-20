import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta de los archivos
archivo = 'julioMaxTimeCodigo.xlsx'
bdVacaciones = 'bdVacaciones.xlsx'
output_file = 'practicas01_filtrado_con_fechas.xlsx'
new_sheet = 'Filtrado_Falso'

# Leer los archivos Excel
dfMaxTime = pd.read_excel(archivo, skiprows=4)
dfVacaciones = pd.read_excel(bdVacaciones)

# Filtramos maxtime
filtro = dfMaxTime[(dfMaxTime['Actividad'] == 'NOV-VACACIONES') & (dfMaxTime['Pais'] == 'Colombia')]

# Ordenar dfVacaciones por Identificacion y alguna columna de fecha (asumiendo que existe)
dfVacaciones = dfVacaciones.sort_values(['Identificacion', 'Fecha_inicio_vacaciones'], ascending=[True, False]).drop_duplicates('Identificacion')

# Realizar el left join
resultado = pd.merge(filtro, dfVacaciones[['Identificacion', 'Fecha_inicio_vacaciones', 'Fecha_fin_vacaciones']], 
                     left_on='Cedula', right_on='Identificacion', how='left')

# Eliminar la columna duplicada de Identificacion si es necesario
resultado = resultado.drop('Identificacion', axis=1)

# Insertar una nueva columna de fecha de los campos separados 
resultado['Reporte_maxtime'] = pd.to_datetime({
    'year': resultado['Año'],
    'month': resultado['Mes'],
    'day': resultado['Dia']
})

# La nueva columna se llamará 'validación' y evaluará la condición lógica
resultado['validacion'] = (resultado['Reporte_maxtime'] >= resultado['Fecha_inicio_vacaciones']) & \
                                  (resultado['Reporte_maxtime'] <= resultado['Fecha_fin_vacaciones'])

# Convertir todas las columnas de tipo datetime al formato corto dd/mm/yyyy
for col in resultado.columns:
    if pd.api.types.is_datetime64_any_dtype(resultado[col]):
        resultado[col] = resultado[col].dt.strftime('%d/%m/%Y')

# Guardar el resultado
resultado.to_excel(output_file, index=False)

# Filtrar filas donde 'validacion' es False
df_falso = resultado[resultado['validacion'] == False]

# Guardar las filas con validacion False en una nueva hoja
with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_falso.to_excel(writer, sheet_name=new_sheet, index=False)

# Aplicar formato en la nueva hoja
wb = load_workbook(output_file)
ws = wb[new_sheet]

# Definir el color amarillo para el formato
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Pintar las celdas con 'False' en amarillo
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.min_column, max_col=ws.max_column):
    for cell in row:
        if cell.value == False:
            cell.fill = yellow_fill

# Guardar el archivo con el formato aplicado
wb.save(output_file)

# Abrir el archivo automáticamente en Windows
os.startfile(output_file)
